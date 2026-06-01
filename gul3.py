import tkinter as tk
import os
from tkinter import messagebox
from tkinter import *
from datetime import datetime
from tkinter import ttk
from openpyxl import Workbook,load_workbook
if not os.path.exists("student.xlsx"):
     wb=Workbook()
     ws=wb.active
     ws.append(["Name","Age","Roll no.","Gendar","Course","Mobile No.","Address"])
     wb.save("student.xlsx")
wb=load_workbook("student.xlsx")
ws=wb.active     
root=tk.Tk()
root.geometry("1500x900")
root.config(bg="white")
root.title("student management system")
headerframe=Frame(root,bd=2,relief="solid",height=1000)
headerframe.pack(fill="both")
lable=tk.Label(headerframe,text="🎓STUDENT MANAGMENT SYSTEM",font=("Arial",35,"bold"))
lable.pack()
dateframe=Frame(headerframe,bg="lightblue",bd=1,height=50,relief="solid",width=200)
dateframe.place(x=1320,y=7)
time_lable=tk.Label(dateframe,text="date and time",bg="lightblue",width=15)
time_lable.pack()
date_lable=tk.Label(dateframe,text="date",bg="lightblue")
date_lable.pack()
def time():
    currentdate=datetime.now().strftime("📆%d %b %y")
    currenttime=datetime.now().strftime("🕐%I:%m:%S %p")
    date_lable.config(text=currentdate,font=("Arial",12))
    time_lable.config(text=currenttime,font=("Arial",12))
    root.after(1000,time)
time()  
leftframe1=Frame(root,bg="lightblue",bd=2,relief="solid",height=500,width=400)
leftframe1.place(x=10,y=100)
lable1=tk.Label(leftframe1,bg="lightblue",text="STUDENT DETAILS",font=("Arial",25,"bold"))
lable1.place(x=40,y=20)
lineframe=Frame(leftframe1,bd=2,bg="black",width=400)
lineframe.place(x=0,y=25)
lineframe2=Frame(leftframe1,bd=2,bg="black",width=400)
lineframe2.place(x=0,y=60)
lable2=tk.Label(leftframe1,text="NAME :",bg="lightblue",font=("Arial",15,"bold"))
lable2.place(x=10,y=70)
entry=tk.Entry(leftframe1,font=("Arial",10),width=23)
entry.place(x=180,y=70,height=30)
entry.insert(0,"Enter the name")
lable3=tk.Label(leftframe1,text="AGE :",bg="lightblue",font=("Arial",15,"bold"))
lable3.place(x=10,y=110)
entry2=tk.Entry(leftframe1,font=("Arial",10),width=23)
entry2.place(x=180,y=110,height=30) 
entry2.insert(0,"Enter the age")
lable4=tk.Label(leftframe1,text="ROLL NO :",bg="lightblue",font=("Arial",15,"bold"))
lable4.place(x=10,y=150)
entry3=tk.Entry(leftframe1,font=("Arial",10),width=23)
entry3.place(x=180,y=150,height=30) 
entry3.insert(0,"Enter the roll no")
lable5=tk.Label(leftframe1,text="GENDAR :",bg="lightblue",font=("Arial",15,"bold"))
lable5.place(x=10,y=190)
la=ttk.Combobox(leftframe1,font=("Arial",10))
la["values"]=("MALE","FEMALE","OTHERS")
la.place(x=180,y=190,height=30)
la.insert(0,"Enter the gendar")
lable6=tk.Label(leftframe1,text="COUSE:",bg="lightblue",font=("Arial",15,"bold"))
lable6.place(x=10,y=230)
la1=ttk.Combobox(leftframe1,font=("Arial",10))
la1["values"]=("BBA","BA","B.TECH","B.COM","MBA")
la1.place(x=180,y=230,height=30)
la1.insert(0,"Enter the couse")
lable7=tk.Label(leftframe1,text="PHONE NO:",bg="lightblue",font=("Arial",15,"bold"))
lable7.place(x=10,y=270)
entry4=tk.Entry(leftframe1,font=("Arial",10),width=23)
entry4.place(x=180,y=270,height=30)
entry4.insert(0,"Enter the phone no")
lable8=tk.Label(leftframe1,text="ADDRESS:",bg="lightblue",font=("Arial",15,"bold"))
lable8.place(x=10,y=310)
entry5=tk.Entry(leftframe1,font=("Arial",10),width=23)
entry5.place(x=180,y=310,height=100)
entry5.insert(0,"Enter the add")
rightframe1=Frame(root,bg="white",bd=2,relief="solid",height=500,width=950)
rightframe1.place(x=500,y=100)
lable0=tk.Label(rightframe1,bg="white",fg="black",text="STUDENT RECORCDS",font=("Arial",25,"bold"))
lable0.place(x=300,y=0)
lineframe1=Frame(rightframe1,bd=2,bg="black",width=950)
lineframe1.place(x=0,y=40)
tableframr=Frame(rightframe1,bd=2,relief="solid")
tableframr.place(x=20,y=50,width=880,height=430)
scroolx=Scrollbar(tableframr,orient=HORIZONTAL)
scrooly=Scrollbar(tableframr,orient=VERTICAL)
def select_data(event):
     seletced=studenttable.focus()
     values=studenttable.item(seletced,'values')  
     entry.delete(0,END) 
     entry.insert(0,values[0])  
     entry2.delete(0,END) 
     entry2.insert(0,values[1])
     entry3.delete(0,END) 
     entry3.insert(0,values[2])
     la.delete(0,END) 
     la.insert(0,values[3])
     la1.delete(0,END) 
     la1.insert(0,values[4])
     entry4.delete(0,END) 
     entry4.insert(0,values[5])
     entry5.delete(0,END) 
     entry5.insert(0,values[6])
studenttable=ttk.Treeview(tableframr,columns=("NAME","AGE","ROLL NO.","GENDAR","COUSE","PHONE NO.","ADDRESS"),xscrollcommand=scroolx.set,yscrollcommand=scrooly.set)
scroolx.pack(side="bottom",fill="x")
scrooly.pack(side="right",fill="y")
scroolx.config(command=studenttable.xview)
scrooly.config(command=studenttable.xview)
studenttable.heading("NAME",text="NAME")
studenttable.heading("AGE",text="AGE")
studenttable.heading("ROLL NO.",text="ROLL NO.")
studenttable.heading("GENDAR",text="GENDAR")
studenttable.heading("COUSE",text="COUSE")
studenttable.heading("PHONE NO.",text="PHONE NO.")
studenttable.heading("ADDRESS",text="ADDRESS")
studenttable.column("NAME",width=100)
studenttable.column("AGE",width=180)
studenttable.column("ROLL NO.",width=80)
studenttable.column("GENDAR",width=120)
studenttable.column("COUSE",width=120)
studenttable.column("PHONE NO.",width=150)
studenttable.column("ADDRESS",width=180)
studenttable["show"]=("headings")
studenttable.pack(fill="both",expand=1)
studenttable.bind("<ButtonRelease-1>",select_data)
def save():
    studenttable.insert(
        '',
        END,
        values=(entry.get(),entry2.get(),entry3.get(),la.get(),la1.get(),entry4.get(),
                entry5.get())                   
    )
    ws.append([entry.get(),entry2.get(),entry3.get(),la.get(),la1.get(),entry4.get(),
                entry5.get()])
    wb.save("student.xlsx")
def clear():
        entry.delete(0,END),entry2.delete(0,END),entry3.delete(0,END),entry4.delete(0,END), entry5.delete(0,END),la.set("selete gendar"),la1.set("selete couse")
def Delete():
     selected_item=studenttable.selection()
     studenttable.delete(selected_item)  
def search():
     search_value=entry3.get()
     for item in studenttable.get_children():          
        data=studenttable.item(item)
        if search_value in str(data["values"]):
            studenttable.selection_set(item)
            studenttable.focus(item)
            studenttable.see(item) 

def update():
    selected_row=studenttable.focus()
    studenttable.item(selected_row,values=(
         entry.get(),
         entry2.get(),
         entry3.get(),
         la.get(),
         la1.get(),
         entry4.get(),
         entry5.get()
         
    ))  
def exit():
     root.destroy()                      
buttonframr=Frame(root,relief="solid",bg="white")
buttonframr.place(x=0,y=700,width=1500,height=50)
button=tk.Button(buttonframr,text="💾SAVE",bg="red",fg="white",font=("Arial",17,"bold"),command=save)
button.place(x=100,y=0)
button1=tk.Button(buttonframr,text="🔎SEARCH",bg="green",fg="white",font=("Arial",17,"bold"),command=search)
button1.place(x=350,y=0)
button2=tk.Button(buttonframr,text="🖊UPDATE",bg="orange",fg="white",font=("Arial",17,"bold"),command=update)
button2.place(x=600,y=0)
button3=tk.Button(buttonframr,text="🗑DELETE",bg="blue",fg="white",font=("Arial",17,"bold"),command=Delete)
button3.place(x=850,y=0)
button4=tk.Button(buttonframr,text="🥍CLEAR",bg="pink",fg="white",font=("Arial",17,"bold"),command=clear)
button4.place(x=1100,y=0)
button5=tk.Button(buttonframr,text="🚪EXIT",bg="gray",fg="white",font=("Arial",17,"bold"),command=exit)
button5.place(x=1350,y=0)
for row in ws.iter_rows(min_row=2,values_only=True):
     studenttable.insert("",END,values=row)
root.mainloop();